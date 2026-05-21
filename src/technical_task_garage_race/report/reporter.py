"""Excel report generation from call transcripts.

The report layer converts transcript files into structured call summaries
and writes them into the workbook template.
"""

from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from technical_task_garage_race.report.constants import (
    DATA_START_ROW,
    HEADER_ROW,
    NAME_RE,
    RED,
    SHEET_NAME,
)
from technical_task_garage_race.report.schema import CallSummary
from technical_task_garage_race.report.utils import (
    extract_date,
    extract_phone,
    normalize_text,
    read_text,
    score_yes_no,
)


def summarize_transcript(transcript_path: Path) -> CallSummary:
    """Convert one transcript into a structured call summary.

    Args:
        transcript_path: Path to the transcript text file.

    Returns:
        CallSummary: Parsed metadata, checklist scores, and comments.
    """
    text = read_text(transcript_path)
    normalized = normalize_text(text)

    phone = extract_phone(transcript_path, text)
    call_type = "Вхідний" if "incoming" in transcript_path.stem else "Вихідний"
    manager_match = NAME_RE.search(text)
    manager = manager_match.group(1) if manager_match else ""

    work_keywords = [
        ("ТО", ["то", "технічн", "техобсл", "обслуговуван"]),
        ("заміна мастил", ["мастил", "олив", "масло", "роздатк", "коробк", "міст"]),
        ("діагностика", ["діагност", "перевір", "огляд", "передпродаж"]),
        (
            "ремонт системи охолодження",
            ["антифриз", "охолодж", "радіатор", "патрубк", "тече"],
        ),
    ]
    work_type = ""
    for label, keywords in work_keywords:
        if any(keyword in normalized for keyword in keywords):
            work_type = label
            break
    if not work_type:
        work_type = "Не визначено"

    scores = {
        "greeting": score_yes_no(normalized, ["доброго дня", "добрий день", "вітаю"]),
        "need_discovery": score_yes_no(
            normalized,
            [
                "який у вас автомобіль",
                "що за автомобіль",
                "під ким",
                "які роботи",
                "що потрібно",
            ],
            ["не знаю"],
        ),
        "schedule": score_yes_no(
            normalized, ["запис", "можу запропонувати", "на", "приїхати"]
        ),
        "closing": score_yes_no(
            normalized, ["дякую", "гарного дня", "до зустрічі", "до зв'язку"]
        ),
    }

    bad_comment = ""
    if any(
        keyword in normalized
        for keyword in [
            "не знаю",
            "не коррект",
            "неправильно",
            "не правильно",
            "не задав",
            "не уточнив",
            "не відповів",
            "не сказав",
        ]
    ):
        bad_comment = "Менеджер відповідав некоректно або не уточнив важливі деталі."

    elif any(
        keyword in normalized
        for keyword in ["перегрів", "ризик", "небезп", "тече", "калюжа"]
    ):
        bad_comment = "Є ризики по стану авто або розмова містить проблемний сигнал по технічному стану."

    result = (
        "Передзвонити"
        if any(
            keyword in normalized
            for keyword in ["передзвон", "ще наберу", "потім", "подзвоню", "відпишу"]
        )
        else "Запис"
    )
    total_score = sum(scores.values())

    date_value = extract_date(transcript_path)
    if not date_value:
        date_value = date.today().isoformat()

    return CallSummary(
        source_path=transcript_path,
        date=date_value,
        call_type=call_type,
        phone=phone,
        branch="",
        manager=manager,
        work_type=work_type,
        result=result,
        comment=bad_comment,
        scores=scores,
        total_score=total_score,
    )


def build_report(
    template_path: Path, transcripts_dir: Path, output_path: Path
) -> list[CallSummary]:
    """Build or extend an Excel report from transcript files.

    Args:
        template_path: Path to the workbook template.
        transcripts_dir: Directory containing transcript text files.
        output_path: Path where the resulting workbook should be saved.

    Returns:
        list[CallSummary]: List of transcript summaries that were written to
        the workbook.

    Raises:
        FileNotFoundError: If the template or transcript directory does not
        exist.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    if not transcripts_dir.exists():
        raise FileNotFoundError(f"Transcripts directory not found: {transcripts_dir}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    created_new_workbook = not output_path.exists()
    if created_new_workbook:
        shutil.copy2(template_path, output_path)
    workbook = load_workbook(output_path)

    worksheet = workbook[SHEET_NAME]

    if created_new_workbook and worksheet.max_row >= DATA_START_ROW:
        worksheet.delete_rows(DATA_START_ROW, worksheet.max_row - DATA_START_ROW + 1)

    worksheet.cell(row=HEADER_ROW, column=21, value="Сума балів")

    transcript_paths = sorted(transcripts_dir.glob("*.txt"))

    summaries: list[CallSummary] = []

    for path in transcript_paths:
        summary = summarize_transcript(path)
        summaries.append(summary)

    start_row = worksheet.max_row + 1 if not created_new_workbook else DATA_START_ROW
    for i, summary in enumerate(summaries):
        row_index = start_row + i

        worksheet.cell(row=row_index, column=1, value=summary.date)
        worksheet.cell(row=row_index, column=2, value=summary.call_type)
        worksheet.cell(row=row_index, column=3, value=summary.phone)
        worksheet.cell(row=row_index, column=4, value=summary.branch)
        worksheet.cell(row=row_index, column=5, value=summary.manager)
        worksheet.cell(row=row_index, column=6, value=summary.scores["greeting"])
        worksheet.cell(row=row_index, column=7, value=summary.scores["need_discovery"])
        worksheet.cell(row=row_index, column=8, value=0)
        worksheet.cell(row=row_index, column=9, value=0)
        worksheet.cell(row=row_index, column=10, value=0)
        worksheet.cell(row=row_index, column=11, value=summary.scores["schedule"])
        worksheet.cell(row=row_index, column=12, value=0)
        worksheet.cell(row=row_index, column=13, value=summary.scores["closing"])
        worksheet.cell(row=row_index, column=14, value=summary.work_type)
        worksheet.cell(
            row=row_index,
            column=15,
            value=1 if summary.work_type != "Не визначено" else 0,
        )
        worksheet.cell(row=row_index, column=16, value="")
        worksheet.cell(row=row_index, column=17, value=summary.result)
        worksheet.cell(row=row_index, column=18, value=summary.total_score)
        worksheet.cell(row=row_index, column=19, value="")
        comment_cell = worksheet.cell(row=row_index, column=20, value=summary.comment)
        worksheet.cell(row=row_index, column=21, value=summary.total_score)

        if summary.comment:
            comment_cell.font = Font(color=RED)

    workbook.save(output_path)
    return summaries
